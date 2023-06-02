import sys, os
import openpyxl as op
from PyQt5 import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5 import uic
import random

#TODO
#StudentList 불러오기(완)
#StudentList filePath 불러오기(완)
#StudentList fileName 불러오기(완)
#MBTI별로 분류하기(완)
#spinBox로 팀원수, 팀수 입력받기(완)
#teambuilding(미완)
#TeamfilePath 불러오기(완)
#exit 활성화(완)

###########################################################

form_class = uic.loadUiType("MBTI_teamBuilding.ui")[0]

class MainWindow(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.initUI()
    
    def initUI(self):
        self.Exit.setStyleSheet("color: red;"
                             "border-style: solid;"
                             "border-width: 2px;"
                             "border-color: #FA8072;"
                             "border-radius: 3px")
        self.loadStudentFile.clicked.connect(self.loadStudentlist)
        self.compareMBTI.clicked.connect(self.compare)
        self.teamCountspin.valueChanged.connect(self.countTeam)
        self.teammateCountspin.valueChanged.connect(self.countTeammate)
        self.BuildTeam.clicked.connect(self.buildTeam)
        self.Exit.clicked.connect(self.exit)
        self.studentList = []
        self.studentE = []
        self.studentI = []
        self.studentETeam = []
        self.studentITeam = []
        self.teammate = []
        self.realTeammate = []
        self.teamNum = 0
        self.teammateNum = 0
        
        
    #학생명단 불러오기
    def loadStudentlist(self):
        path = QFileDialog.getOpenFileName(self, 'Open File', '', 'xlsx File(*.xlsx)')[0]
        
        if path == '':
                QMessageBox.warning(self, "경고", "파일경로를 읽어올 수 없습니다")
                return
            
        # 파일경로/파일명 분리
        dirPath, Name = os.path.split(path)
        self.Name.insertPlainText(Name)
        self.filePath.insertPlainText(dirPath)

        #위 절대 경로 활용해 openpyxl workbook 객체 생성
        wb = op.load_workbook(path)
        sheet = wb.active
            
        #excel 파일의 행과 열 입력
        maxCol = sheet.max_column
        maxRow = sheet.max_row
        sheetData = list(sheet.values)

        self.studentData.setRowCount(maxRow-1)
        self.studentData.setColumnCount(maxCol)
        self.studentData.setHorizontalHeaderLabels(sheetData[0])
        
        row_index = 0
        for value_tuple in sheetData[1:]:
            self.studentList.append(list(value_tuple))
            col_index = 0
            for value in value_tuple:
                if value==None:
                    value = ''
                self.studentData.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1
    
    #MBTI 분류하기
    def compare(self):
        if self.studentList == []:
            QMessageBox.critical(self, '주의', '학생명단이 존재하지 않습니다')
            return
        
        dataMaxrow = len(self.studentList)
        
        for Row in range(dataMaxrow):
            if self.studentList[Row][2] == 'E':
                self.studentE.append(self.studentList[Row][0:2])
                
            
            elif self.studentList[Row][2] == 'I':
                self.studentI.append(self.studentList[Row][0:2])
                # print(self.studentI)
            
            
            else:
                QMessageBox.information(self, '알림', 'MBTI에 오류가 있습니다')
            
        QMessageBox.information(self, '알림', '분류가 완료되었습니다')
        # print(self.studentE)
        # print('--------------')
        # print(self.studentI)
        
    #팀 개수 입력받기
    def countTeam(self):
        self.teamNum = self.teamCountspin.value()
        
    #팀원 수 입력받기
    def countTeammate(self):
        self.teammateNum = self.teammateCountspin.value()
    
    # #팀 만들기
    def buildTeam(self):
        # print(type(self.teamNum))
        # print(type(self.teammateNum))
        for i in range(self.teamNum):
            if int(self.teammateNum%2) == 0:
                
                self.studentETeam = random.sample(self.studentE, int(self.teammateNum/2))
                self.studentITeam = random.sample(self.studentI, int(self.teammateNum/2))
                self.teammate.append(self.studentETeam)
                self.teammate.append(self.studentITeam)
                
                #추출된 인원 리스트에서 삭제
                if self.studentE == self.studentETeam:
                    self.studentETeam.remove(self.studentE)
                
                #추출된 인원 리스트에서 삭제
                if self.studentI == self.studentITeam:
                    self.studentITeam.remove(self.studentI)

                #3차원 리스트를 2차원 리스트로 변환(리스트 컴프리헨션)
                self.realTeammate = [inner_list for outer_list in self.teammate for inner_list in outer_list]
                # print('팀원 리스트:', self.realTeammate)
                
                wb = op.Workbook()
                sheet = wb.active
                sheet.append(["이름", "학번"])
                filename = f'team{i+1}.xlsx'
                print(filename)
                
                #엑셀에 데이터 입력
                for Rows in range(self.teammateNum):
                    sheet.append([self.realTeammate[Rows][0], str(self.realTeammate[Rows][1])])
                
                #엑셀 저장
                wb.save(filename)
                
                #내보내기 파일 화면 표시
                self.TeamPath.appendPlainText(os.getcwd() + filename)
                
            
            elif self.teammateNum%2 == 1:
                self.studentETeam = random.sample(self.studentE, int(self.teammateNum/2+1))
                self.studentITeam = random.sample(self.studentI, int(self.teammateNum/2))
                self.teammate.append(self.studentETeam)
                self.teammate.append(self.studentITeam)
                
                #추출된 인원 리스트에서 삭제
                if self.studentE == self.studentETeam:
                    self.studentE.remove(self.studentETeam)
                
                #추출된 인원 리스트에서 삭제
                if self.studentI == self.studentITeam:
                    self.studentI.remove(self.studentITeam)

                #3차원 리스트를 2차원 리스트로 변환(리스트 컴프리헨션)
                self.realTeammate = [inner_list for outer_list in self.teammate for inner_list in outer_list]
                # print('팀원 리스트:', self.realTeammate)
                
                wb = op.Workbook()
                sheet = wb.active
                sheet.append(["이름", "학번"])
                filename = f'team{i+1}.xlsx'
                print(filename)
                
                #엑셀에 데이터 입력
                for Rows in range(self.teammateNum):
                    sheet.append([self.realTeammate[Rows][0], str(self.realTeammate[Rows][1])])
                
                #엑셀 저장
                wb.save(filename)
                
                #내보내기 파일 화면 표시
                self.TeamPath.appendPlainText(os.getcwd() + filename)
        
           

    def exit(self):
        re = QMessageBox.question(self, "확인", "프로그램을 종료 하시겠습니까?", QMessageBox.Yes|QMessageBox.No)

        if re == QMessageBox.Yes:
            QApplication.quit()
    
if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #MainWindow 인스턴스 생성
    myWindow = MainWindow() 

    #프로그램 화면을 보여주는 코드
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    sys.exit(app.exec_())